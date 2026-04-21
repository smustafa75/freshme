"""
Builds a Cloud_Services_Update_YYYY-MM-DD.xlsx from live AWS + OCI RSS/Atom feeds.
Sheets: Summary | Events | AWS Updates | OCI Updates
Events are scraped dynamically from official event pages; OCI falls back to
public Oracle RSS feeds if per-service docs feeds are blocked.
"""
from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from email.utils import parsedate_to_datetime
from urllib.request import Request, urlopen
from xml.etree import ElementTree as ET
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Runtime constants ─────────────────────────────────────────────────────────
UAE_TZ        = ZoneInfo("Asia/Dubai")
LOOKBACK_HOURS = 72   # covers the 3-day run interval; raise to 96 for extra buffer
HTTP_TIMEOUT   = 20
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.6367.207 Safari/537.36"
)

# ── AWS feed sources ──────────────────────────────────────────────────────────
AWS_FEEDS = [
    ("AWS What's New", "https://aws.amazon.com/about-aws/whats-new/recent/feed/"),
    ("AWS News Blog",  "https://aws.amazon.com/blogs/aws/feed/"),
]

# ── OCI per-service release-note feeds ───────────────────────────────────────
# URL pattern: https://docs.oracle.com/en-us/iaas/releasenotes/services/<slug>/feed
OCI_SERVICE_FEEDS = [
    ("Compute",                         "compute"),
    ("Object Storage",                  "object-storage"),
    ("Networking",                      "networking"),
    ("Autonomous Database (Shared)",    "autonomous-database-shared"),
    ("Autonomous Database (Dedicated)", "autonomous-database-dedicated"),
    ("Base Database",                   "base-database"),
    ("Kubernetes Engine (OKE)",         "kubernetes-engine"),
    ("Functions",                       "functions"),
    ("Generative AI",                   "generative-ai"),
    ("Generative AI Agents",            "generative-ai-agents"),
    ("Data Science",                    "data-science"),
    ("Identity & Access Management",    "identity-and-access-management"),
    ("Monitoring",                      "monitoring"),
    ("Logging",                         "logging"),
    ("FastConnect",                     "fastconnect"),
    ("Load Balancer",                   "load-balancer"),
    ("Database Migration",              "database-migration"),
    ("GoldenGate",                      "goldengate"),
]

# Fallback RSS sources used when all per-service feeds are blocked/empty.
OCI_FALLBACK_FEEDS = [
    ("OCI Release Notes", "https://docs.oracle.com/en-us/iaas/releasenotes/services/feed"),
    ("Oracle Blog",       "https://blogs.oracle.com/cloud-infrastructure/rss"),
    ("Oracle Tech Blog",  "https://blogs.oracle.com/developers/rss"),
]

# ── Event pages (dates scraped dynamically from static HTML) ──────────────────
# Add/remove entries here when AWS or Oracle rename/cancel a flagship event.
# Summits and rolling tours are omitted — their pages are JS-rendered.
_EVENT_PAGES = [
    {"name": "AWS re:Invent",  "provider": "AWS",
     "url": "https://aws.amazon.com/events/reinvent/",
     "notes": "AWS's flagship annual conference — compute, AI, data, and security launches."},
    {"name": "Oracle AI World", "provider": "Oracle",
     "url": "https://www.oracle.com/cloudworld/",
     "notes": "Oracle's flagship annual conference focused on AI, database, and cloud."},
]

# month name/abbrev → int (used by the date-scraping regex)
_MONTH_MAP = {m: i for i, m in enumerate(
    ["january","february","march","april","may","june",
     "july","august","september","october","november","december"], 1)}
_MONTH_MAP.update({k[:3]: v for k, v in list(_MONTH_MAP.items())})

# Regex: matches "November 30 - December 4, 2026 | Las Vegas, NV" and variants
_DATE_RE = re.compile(
    r"(January|February|March|April|May|June|July|August|"
    r"September|October|November|December)\s+(\d{1,2})"
    r"(?:\s*[-–]\s*(?:(January|February|March|April|May|June|July|August|"
    r"September|October|November|December)\s+)?(\d{1,2}))?"
    r",?\s+(\d{4})(?:\s*[|,]\s*([^\n<\[]{3,40}))?",
    re.IGNORECASE,
)

# ── Data model ────────────────────────────────────────────────────────────────
@dataclass
class Item:
    title:     str
    link:      str
    published: datetime   # always tz-aware UTC
    summary:   str = ""
    source:    str = ""
    category:  str = ""

# ── HTTP helpers ──────────────────────────────────────────────────────────────
def _fetch(url: str) -> bytes:
    """GET url with a browser User-Agent; transparently decompresses gzip."""
    req = Request(url, headers={
        "User-Agent":      USER_AGENT,
        "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection":      "keep-alive",
        "Referer":         "https://www.google.com/",
        "DNT":             "1",
    })
    with urlopen(req, timeout=HTTP_TIMEOUT) as r:
        data = r.read()
        if r.info().get("Content-Encoding") == "gzip":
            import gzip
            data = gzip.decompress(data)
        return data


def _strip_html(s: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", s or "")).strip()


def _excerpt(text: str, n: int = 2, max_chars: int = 280) -> str:
    """Return first n sentences of plain text, capped at max_chars."""
    text = _strip_html(text)
    if not text:
        return ""
    out = " ".join(re.split(r"(?<=[.!?])\s+", text)[:n]).strip()
    return (out[:max_chars - 1] + "…") if len(out) > max_chars else out


def _parse_date(s: str) -> datetime | None:
    """Parse RFC 2822 or ISO 8601 date string → UTC-aware datetime."""
    if not s:
        return None
    try:
        dt = parsedate_to_datetime(s)
    except (TypeError, ValueError):
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        except ValueError:
            return None
    if dt and dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc) if dt else None

# ── Feed parsing ──────────────────────────────────────────────────────────────
_NS = {
    "atom":    "http://www.w3.org/2005/Atom",
    "content": "http://purl.org/rss/1.0/modules/content/",
    "dc":      "http://purl.org/dc/elements/1.1/",
}

def parse_feed(source: str, xml_bytes: bytes) -> list[Item]:
    """Parse RSS 2.0 or Atom 1.0 bytes into a list of Items."""
    items: list[Item] = []
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return items

    # RSS 2.0 — <item> elements
    for it in root.findall(".//item"):
        title    = (it.findtext("title") or "").strip()
        link     = (it.findtext("link") or "").strip()
        pub      = _parse_date(it.findtext("pubDate") or it.findtext("dc:date", default="", namespaces=_NS))
        summary  = it.findtext("description") or it.findtext("content:encoded", default="", namespaces=_NS) or ""
        category = (it.findtext("category") or "").strip()
        if title and link and pub:
            items.append(Item(title, link, pub, _excerpt(summary), source, category))

    # Atom 1.0 — <entry> elements (only if RSS found nothing)
    if not items:
        for e in root.findall("atom:entry", _NS):
            title   = (e.findtext("atom:title", default="", namespaces=_NS) or "").strip()
            lel     = e.find("atom:link[@rel='alternate']", _NS) or e.find("atom:link", _NS)
            link    = (lel.get("href") or "").strip() if lel is not None else ""
            pub     = _parse_date(
                e.findtext("atom:updated",   default="", namespaces=_NS) or
                e.findtext("atom:published", default="", namespaces=_NS))
            summary = (e.findtext("atom:summary", default="", namespaces=_NS) or
                       e.findtext("atom:content", default="", namespaces=_NS) or "")
            cats    = [c.get("term", "") for c in e.findall("atom:category", _NS) if c.get("term")]
            if title and link and pub:
                items.append(Item(title, link, pub, _excerpt(summary), source, ", ".join(cats)[:80]))
    return items


def fetch_feed(source: str, url: str) -> list[Item]:
    try:
        return parse_feed(source, _fetch(url))
    except Exception as ex:
        print(f"  [warn] {source}: {ex}", file=sys.stderr)
        return []

# ── AWS service categorisation ────────────────────────────────────────────────
# (display_name, regex) — first match wins against the announcement title.
AWS_SERVICE_PATTERNS = [
    ("Amazon EC2",          r"\bamazon ec2\b|\bec2\b"),
    ("AWS Lambda",          r"\blambda\b"),
    ("Amazon S3",           r"\bamazon s3\b|\bs3\b"),
    ("Amazon ECS",          r"\becs\b"),
    ("Amazon EKS",          r"\beks\b"),
    ("Amazon CloudFront",   r"cloudfront"),
    ("Amazon Bedrock",      r"bedrock"),
    ("Amazon CloudWatch",   r"cloudwatch"),
    ("Amazon VPC",          r"\bvpc\b"),
    ("Amazon RDS",          r"\brds\b"),
    ("Amazon DynamoDB",     r"dynamodb"),
    ("Amazon SageMaker",    r"sagemaker"),
    ("Amazon OpenSearch",   r"opensearch"),
    ("Amazon Aurora",       r"aurora"),
    ("Amazon SNS",          r"\bsns\b"),
    ("Amazon SQS",          r"\bsqs\b"),
    ("Amazon Kinesis",      r"kinesis"),
    ("Amazon EMR",          r"\bemr\b"),
    ("AWS IAM",             r"\biam\b|identity and access management"),
    ("AWS Glue",            r"\bglue\b"),
    ("AWS Step Functions",  r"step functions"),
    ("Amazon API Gateway",  r"api gateway"),
    ("AWS CloudFormation",  r"cloudformation"),
    ("AWS Transform",       r"aws transform"),
    ("AWS Network Firewall",r"network firewall"),
    ("AWS Security Hub",    r"security hub"),
    ("Amazon GuardDuty",    r"guardduty"),
]

def guess_aws_service(title: str) -> str:
    """Return the best-matching AWS service name for an announcement title."""
    t = title.lower()
    for name, pat in AWS_SERVICE_PATTERNS:
        if re.search(pat, t):
            return name
    m = re.search(r"\b(Amazon|AWS)\s+([A-Z][A-Za-z0-9]+(?:\s+[A-Z][A-Za-z0-9]+){0,2})", title)
    return m.group(0) if m else "AWS"

# ── Event scraping & filtering ────────────────────────────────────────────────
def _parse_event_page(meta: dict) -> dict | None:
    """Fetch an event page and extract the next future start/end date + location."""
    try:
        html = _fetch(meta["url"]).decode("utf-8", errors="ignore")
    except Exception as ex:
        print(f"  [warn] Could not fetch {meta['name']} page: {ex}", file=sys.stderr)
        return None

    year_now = datetime.now().year
    for m in _DATE_RE.finditer(html):
        year = int(m.group(5))
        if year < year_now:
            continue
        sm = _MONTH_MAP[m.group(1).lower()];  sd = int(m.group(2))
        em = _MONTH_MAP[m.group(3).lower()] if m.group(3) else sm
        ed = int(m.group(4)) if m.group(4) else sd
        loc = re.split(r'["{\\]', (m.group(6) or "").strip().rstrip(".,"))[0].strip() \
              or meta.get("location", "TBA")
        return {**meta,
                "start": f"{year}-{sm:02d}-{sd:02d}",
                "end":   f"{year}-{em:02d}-{ed:02d}",
                "location": loc}

    print(f"  [warn] No future date found on {meta['name']} page", file=sys.stderr)
    return None


def fetch_events() -> list[dict]:
    """Scrape dates from official event pages; skip entries with no future date."""
    return [e for e in (_parse_event_page(m) for m in _EVENT_PAGES) if e]


def active_or_upcoming_events(today_uae: datetime, raw_events: list[dict],
                               window_days: int = 30) -> list[dict]:
    """Filter raw events to those that are LIVE or start within window_days."""
    out, today = [], today_uae.date()
    for e in raw_events:
        if not e.get("start"):
            continue
        start = datetime.fromisoformat(e["start"]).date()
        end   = datetime.fromisoformat(e["end"]).date()
        if start <= today <= end:
            status, days_to = "LIVE", 0
        elif today < start <= today + timedelta(days=window_days):
            status, days_to = "UPCOMING", (start - today).days
        else:
            continue
        out.append({**e, "status": status, "days_to": days_to,
                    "start_date": start, "end_date": end})
    out.sort(key=lambda x: (x["status"] != "LIVE", x["days_to"]))
    return out


def items_for_event(items: list[Item], event_name: str) -> list[Item]:
    """Return feed items whose title+summary mention the event (fuzzy match)."""
    key   = event_name.lower()
    short = re.sub(r"[^a-z0-9]", "", key)
    return [it for it in items
            if key in (it.title + " " + it.summary).lower()
            or short in re.sub(r"[^a-z0-9]", "", (it.title + it.summary).lower())]

# ── Excel style constants ─────────────────────────────────────────────────────
_THIN   = Side(border_style="thin", color="D0D0D0")
BORDER  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
F_HEAD  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
F_BODY  = Font(name="Arial", size=10)
F_LINK  = Font(name="Arial", size=10, color="0563C1", underline="single")
F_TITLE = Font(name="Arial", bold=True, size=14)
F_META  = Font(name="Arial", size=9, italic=True, color="666666")

def _hdr(ws, row: int, ncols: int, value: str, font: Font) -> None:
    """Write a merged header row spanning ncols columns."""
    ws.cell(row=row, column=1, value=value).font = font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

def _fill(ws, row: int, col: int, value, font: Font, fill_color: str | None = None,
          merge_end_col: int | None = None, wrap: bool = False, height: int | None = None):
    """Write a cell with font, optional fill, optional merge, optional wrap."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill_color:
        c.fill = PatternFill("solid", start_color=fill_color)
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical="top")
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
    if height:
        ws.row_dimensions[row].height = height
    return c

# ── Excel sheet builders ──────────────────────────────────────────────────────
def _updates_sheet(ws, rows: list, title: str, fill_color: str, timestamp: str) -> None:
    """Populate an AWS or OCI updates sheet."""
    _hdr(ws, 1, 5, title, F_TITLE)
    _hdr(ws, 2, 5, timestamp, F_META)

    # Column headers
    for col, h in enumerate(["Service", "Category", "What's Added / Updated", "Source", "Link"], 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font  = F_HEAD
        c.fill  = PatternFill("solid", start_color=fill_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    if not rows:
        c = ws.cell(row=5, column=1, value="No new announcements captured in the lookback window.")
        c.font = Font(name="Arial", italic=True, color="888888")
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=5)
    else:
        for i, row in enumerate(rows, start=5):
            for col, val in enumerate(row, 1):
                c = ws.cell(row=i, column=col, value=val)
                c.font = F_BODY;  c.border = BORDER
                c.alignment = Alignment(vertical="top", wrap_text=True)
            # Make the link cell clickable
            if row[4]:
                lc = ws.cell(row=i, column=5)
                lc.hyperlink = row[4];  lc.font = F_LINK

    for col, w in {"A": 26, "B": 22, "C": 70, "D": 18, "E": 50}.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[4].height = 24
    for r in range(5, 5 + max(len(rows), 1)):
        ws.row_dimensions[r].height = 50
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


def _events_sheet(ws, events: list, aws_items: list, oci_items: list, timestamp: str) -> None:
    """Populate the Events sheet with LIVE/UPCOMING event blocks."""
    _hdr(ws, 1, 5, "Cloud Events — Daily Summary", F_TITLE)
    _hdr(ws, 2, 5, timestamp, F_META)

    if not events:
        c = ws.cell(row=4, column=1, value="No live or upcoming events in the next 30 days.")
        c.font = Font(name="Arial", italic=True, color="888888")
        ws.column_dimensions["A"].width = 60
        ws.sheet_view.showGridLines = False
        return

    row = 4
    for e in events:
        # Coloured banner: green = LIVE, amber = UPCOMING
        color = "107C10" if e["status"] == "LIVE" else "8A6D00"
        _fill(ws, row, 1, f"[{e['status']}]   {e['name']}   —   {e['provider']}",
              Font(name="Arial", bold=True, size=12, color="FFFFFF"),
              fill_color=color, merge_end_col=5, height=22)
        row += 1

        # Date / location / countdown line
        timing = "Happening now" if e["status"] == "LIVE" else f"Starts in {e['days_to']} day(s)"
        _fill(ws, row, 1,
              f"Dates: {e['start_date']} → {e['end_date']}   |   Location: {e['location']}   |   {timing}",
              Font(name="Arial", size=10, italic=True, color="333333"), merge_end_col=5)
        row += 1

        # Event description
        _fill(ws, row, 1, e["notes"], F_BODY, merge_end_col=5, wrap=True, height=32)
        row += 1

        # Feed items that mention this event
        pool    = aws_items if e["provider"] == "AWS" else oci_items
        matches = items_for_event(pool, e["name"])
        _fill(ws, row, 1,
              f"Announcements in today's feed referencing this event: {len(matches)}",
              Font(name="Arial", bold=True, size=10), merge_end_col=5)
        row += 1

        if matches:
            for it in matches[:10]:
                _fill(ws, row, 1, f"• {it.title}", F_BODY, merge_end_col=4, wrap=True)
                lc = ws.cell(row=row, column=5, value="Open")
                lc.hyperlink = it.link;  lc.font = F_LINK
                row += 1
        else:
            _fill(ws, row, 1, "(No event-tagged items in today's lookback window.)",
                  Font(name="Arial", italic=True, color="888888"), merge_end_col=5)
            row += 1

        row += 1  # blank spacer between events

    for col, w in {"A": 40, "B": 18, "C": 30, "D": 30, "E": 14}.items():
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False


def _summary_sheet(ws, aws_count: int, oci_count: int, event_count: int, timestamp: str) -> None:
    """Populate the Summary sheet with counts and source links."""
    _hdr(ws, 1, 2, "Daily Cloud Services & Offerings Update",
         Font(name="Arial", bold=True, size=16))
    _hdr(ws, 2, 2, timestamp, Font(name="Arial", size=11, italic=True))

    rows = [
        ("", ""),
        ("Report Contents", ""),
        ("AWS announcements",      f"{aws_count} items in last {LOOKBACK_HOURS}h"),
        ("OCI announcements",      f"{oci_count} items in last {LOOKBACK_HOURS}h"),
        ("Live / upcoming events", f"{event_count} tracked"),
        ("", ""),
        ("Primary Sources", ""),
        ("AWS What's New",    "https://aws.amazon.com/about-aws/whats-new/recent/feed/"),
        ("AWS News Blog",     "https://aws.amazon.com/blogs/aws/feed/"),
        ("OCI Release Notes", "https://docs.oracle.com/en-us/iaas/releasenotes/"),
        ("OCI Blog",          OCI_FALLBACK_FEEDS[0][1]),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        a = ws.cell(row=i, column=1, value=k)
        b = ws.cell(row=i, column=2, value=v)
        if k in ("Report Contents", "Primary Sources"):
            # Section header — full-width blue band
            for c in (a, b):
                c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
                c.fill = PatternFill("solid", start_color="2F5496")
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        else:
            a.font = Font(name="Arial", bold=True, size=10)
            b.font = F_LINK if str(v).startswith("http") else F_BODY
            if str(v).startswith("http"):
                b.hyperlink = v

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 75
    ws.sheet_view.showGridLines = False

# ── Data collection ───────────────────────────────────────────────────────────
def _collect(feeds: list[tuple], cutoff: datetime, tag_category: bool = False) -> list[Item]:
    """Fetch a list of (source, url) feeds, filter by cutoff, sort newest-first."""
    items: list[Item] = []
    for source, url in feeds:
        fetched = fetch_feed(source, url)
        if tag_category:
            for it in fetched:
                it.category = source   # label OCI items with their service name
        items.extend(fetched)
    items = [i for i in items if i.published >= cutoff]
    items.sort(key=lambda x: x.published, reverse=True)
    return items


def collect_aws(cutoff: datetime) -> list[Item]:
    return _collect(AWS_FEEDS, cutoff)


def collect_oci(cutoff: datetime) -> list[Item]:
    """Try per-service feeds first; fall back to public Oracle RSS if all are empty."""
    oci_feeds = [(name, f"https://docs.oracle.com/en-us/iaas/releasenotes/services/{slug}/feed")
                 for name, slug in OCI_SERVICE_FEEDS]
    items = _collect(oci_feeds, cutoff, tag_category=True)

    if not items:
        # Per-service feeds returned nothing (likely blocked) — use fallback sources
        print("  [warn] OCI per-service feeds empty — using fallback feeds", file=sys.stderr)
        items = _collect(OCI_FALLBACK_FEEDS, cutoff, tag_category=True)
    else:
        # Append the aggregate release-notes feed alongside per-service results
        items += _collect([OCI_FALLBACK_FEEDS[0]], cutoff)
        items.sort(key=lambda x: x.published, reverse=True)

    return items

# ── Row formatters ────────────────────────────────────────────────────────────
def _aws_rows(items: list[Item]) -> list[tuple]:
    return [(guess_aws_service(it.title), it.category or "",
             it.summary or it.title, it.source, it.link) for it in items]


def _oci_rows(items: list[Item]) -> list[tuple]:
    return [(it.category or "OCI", "",
             (it.title + (". " + it.summary if it.summary else "")).strip(". "),
             it.source, it.link) for it in items]

# ── Workbook assembly ─────────────────────────────────────────────────────────
def build_workbook(out_path: str) -> str:
    now_uae    = datetime.now(UAE_TZ)
    cutoff_utc = datetime.now(timezone.utc) - timedelta(hours=LOOKBACK_HOURS)
    timestamp  = f"Generated {now_uae.strftime('%A, %d %B %Y, %H:%M %Z')} — lookback {LOOKBACK_HOURS}h"
    date_str   = now_uae.strftime("%d %b %Y")

    print("Fetching event dates from official pages…")
    events = active_or_upcoming_events(now_uae, fetch_events())
    print(f"  {len(events)} event(s) in scope")

    print("Fetching AWS feeds…")
    aws_items = collect_aws(cutoff_utc)
    print(f"  {len(aws_items)} AWS items")

    print("Fetching OCI feeds…")
    oci_items = collect_oci(cutoff_utc)
    print(f"  {len(oci_items)} OCI items")

    wb = Workbook()
    summary_ws = wb.active;  summary_ws.title = "Summary"
    _summary_sheet(summary_ws, len(aws_items), len(oci_items), len(events), timestamp)
    _events_sheet(wb.create_sheet("Events"), events, aws_items, oci_items, timestamp)
    _updates_sheet(wb.create_sheet("AWS Updates"), _aws_rows(aws_items),
                   f"AWS — Daily Services & Offerings Update ({date_str})", "FF9900", timestamp)
    _updates_sheet(wb.create_sheet("OCI Updates"), _oci_rows(oci_items),
                   f"OCI — Daily Services & Offerings Update ({date_str})", "C74634", timestamp)

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    now_uae = datetime.now(UAE_TZ)
    out = f"Cloud_Services_Update_{now_uae.strftime('%Y-%m-%d')}.xlsx"
    build_workbook(out)
    print(f"Saved: {out}")
